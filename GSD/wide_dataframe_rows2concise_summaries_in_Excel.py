# This is meant to use with `uv` to run. 
# First install `uv` with `pip install uv` then run `!uv run this_script.py my_pickled_dataframe_file.pkl`
#-------------------------------------------------------------#
# Pickled dataframe saved as `'raw_complexes_pickled_df.pkl'`.
#-------------------------------------------------------------#
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "numpy",
#   "pandas",
#   "openpyxl",
# ]
# ///

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def create_summary_block(ws, data, start_row, start_col):
    """
    Create a summary block starting at the specified row and column
    """
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Summary headers (row 1 of block)
    headers = ['ID', 'source_tissue', 'total_reads']
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Summary values (row 2 of block)
    values = [data['ID'], data['source_tissue'], data['total_reads']]
    for i, value in enumerate(values):
        cell = ws.cell(row=start_row + 1, column=start_col + i)
        cell.value = value
        cell.alignment = center_align
    
    # Transcript details header (row 4 of block - leave a blank row)
    detail_headers = ['transcript', 'TPM']
    for i, header in enumerate(detail_headers):
        cell = ws.cell(row=start_row + 3, column=start_col + i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = center_align
    
    # Transcript details (rows 5+ of block)
    for j, transcript_info in enumerate(data['transcript_data']):
        detail_row = start_row + 4 + j
        # Transcript name
        ws.cell(row=detail_row, column=start_col).value = transcript_info['transcript']
        # TPM value
        ws.cell(row=detail_row, column=start_col + 1).value = transcript_info['TPM']

def process_dataframe_to_blocks(df):
    """
    Convert tidy dataframe to list of summary block dictionaries
    """
    blocks = []
    
    # Group by unique_grouping_tag to get each block
    for group_tag, group_df in df.groupby('unique_grouping_tag'):
        # Get the common information (should be same for all rows in group)
        first_row = group_df.iloc[0]
        
        # Create transcript data list from all rows in group
        transcript_data = []
        for _, row in group_df.iterrows():
            transcript_data.append({
                'transcript': row['common_nom'],
                'TPM': row['TPM']
            })
        
        # Create block dictionary
        block = {
            'ID': first_row['ID'],
            'source_tissue': first_row['source'],
            'total_reads': first_row['ttl_reads'],
            'transcript_data': transcript_data
        }
        
        blocks.append(block)
    
    return blocks

def create_excel_summary_from_df(df, filename="summary_layout_from_df.xlsx"):
    """
    Create Excel file with 4x3 grid of summary blocks from tidy dataframe
    """
    # Convert dataframe to block format
    data_list = process_dataframe_to_blocks(df)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Summary"
    
    # Configuration for 3 columns, 4 rows layout
    cols_per_page = 3
    block_width = 4  # columns per block (3 data + 1 spacing)
    block_height = 8  # rows per block (2 summary + 1 blank + 1 header + 3 details + 1 spacing)
    
    # Create blocks in 3x4 grid
    for idx, data in enumerate(data_list[:12]):  # Limit to 12 samples (3x4)
        # Calculate position in grid
        grid_col = idx % cols_per_page
        grid_row = idx // cols_per_page
        
        # Calculate actual Excel position
        start_col = 1 + (grid_col * block_width)
        start_row = 1 + (grid_row * block_height)
        
        # Create the block
        create_summary_block(ws, data, start_row, start_col)
    
    # Auto-adjust column widths
    for col in range(1, cols_per_page * block_width + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Save the file
    wb.save(filename)
    print(f"Excel file saved as: {filename}")
    print(f"Created {len(data_list)} summary blocks")
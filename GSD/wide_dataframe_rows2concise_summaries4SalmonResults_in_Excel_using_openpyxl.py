# This is meant to use with `uv` to run. 
# First install `uv` with `pip install uv` then run `!uv run {script_url} {pickle_file_name} {output_name_prefix}` where defined those variables prior
#-------------------------------------------------------------#
# Pickled dataframe saved as `'raw_complexes_pickled_df.pkl'`.
#-------------------------------------------------------------#
# /// script
# requires-python = ">=3.12"
# dependencies = [
#   "numpy",
#   "pandas",
#   "openpyxl",
# ]
# ///

def make_multi_sheet_output_fn(output_name_prefix):
    output_fn = f"{output_name_prefix}_multi_sheet.xlsx"
    return output_fn
def make_single_sheet_output_fn(output_name_prefix):
    output_fn = f"{output_name_prefix}_single_sheet.xlsx"
    return output_fn

def create_summary_block(ws, data, start_row, start_col):
    """
    Create a summary block starting at the specified row and column
    """
    # Header styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    # Pale yellow fill for transcript details header
    transcript_header_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")
    # Silvery fill for transcript details header
    transcript_header_fill = PatternFill(start_color="dee5ee", end_color="dee5ee", fill_type="solid")
    center_align = Alignment(horizontal="center", vertical="center")
    
    # Summary headers (row 1 of block)
    headers = ['ID', 'source_tissue', 'total_reads']
    for i, header in enumerate(headers):
        cell = ws.cell(row=start_row, column=start_col + i)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
    
    # Summary values (row 2 of block)
    values = [data['ID'], data['source_tissue'], data['total_reads']]
    for i, value in enumerate(values):
        cell = ws.cell(row=start_row + 1, column=start_col + i)
        cell.value = value
        cell.alignment = center_align
    
    # Transcript details header (row 3 of block - no blank row)
    detail_headers = ['transcript', 'TPM', 'NumReads']
    #detail_headers = ['transcript', 'TPM'] # if prefer not to have 'NumReads' 
    # included uncomment this; need to also make sure don't add value below
    for i, header in enumerate(detail_headers):
        cell = ws.cell(row=start_row + 2, column=start_col + i)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = transcript_header_fill
        cell.alignment = center_align
    
    # Transcript details (rows 4+ of block)
    for j, transcript_info in enumerate(data['transcript_data']):
        detail_row = start_row + 3 + j
        # Transcript name
        ws.cell(row=detail_row, column=start_col).value = transcript_info['transcript']
        # TPM value
        ws.cell(row=detail_row, column=start_col + 1).value = transcript_info['TPM']
        # NumReads value
        ws.cell(row=detail_row, column=start_col + 2).value = transcript_info['NumReads'] # if prefer not to have 'NumReads' included comment this 
        # out; plus, need to adjust header list above

def process_dataframe_to_blocks(df):
    """
    Convert tidy dataframe to list of summary block dictionaries
    """
    blocks = []
    
    # Group by unique_grouping_tag to get each block
    for group_tag, group_df in df.groupby('unique_grouping_tag'):
        # Get the common information (should be same for all rows in group)
        first_row = group_df.iloc[0]
        
        # Create transcript data list from all rows in group
        transcript_data = []
        for _, row in group_df.iterrows():
            transcript_data.append({
                'transcript': row['common_nom'],
                'TPM': row['TPM'],
                'NumReads': row['NumReads']
            })
        
        # Create block dictionary
        block = {
            'ID': first_row['ID'],
            'source_tissue': first_row['source'],
            'total_reads': first_row['ttl_reads'],
            'transcript_data': transcript_data
        }
        
        blocks.append(block)
    
    return blocks

def create_excel_summary_from_df(df, filename="summary_layout_from_df.xlsx", blocks_per_sheet=48, cols_per_row=3):
    """
    Create Excel file with summary blocks from tidy dataframe
    
    Parameters:
    - df: Input dataframe
    - filename: Output Excel filename
    - blocks_per_sheet: Maximum blocks per sheet (default 48 = 16 rows x 3 cols)
    - cols_per_row: Number of columns per row (default 3)
    """
    # Convert dataframe to block format
    data_list = process_dataframe_to_blocks(df)
    total_blocks = len(data_list)
    
    wb = Workbook()
    
    # Configuration
    block_width = 4  # columns per block (3 data + 1 spacing)
    block_height = 7  # rows per block (2 summary + 1 header + 3 details + 1 spacing)
    
    # Calculate number of sheets needed
    num_sheets = (total_blocks + blocks_per_sheet - 1) // blocks_per_sheet
    
    for sheet_num in range(num_sheets):
        # Create or use worksheet
        if sheet_num == 0:
            ws = wb.active
            ws.title = f"Summary_Page_{sheet_num + 1}_of_{num_sheets}"
        else:
            ws = wb.create_sheet(f"Summary_Page_{sheet_num + 1}_of_{num_sheets}")
        
        # Calculate blocks for this sheet
        start_idx = sheet_num * blocks_per_sheet
        end_idx = min(start_idx + blocks_per_sheet, total_blocks)
        sheet_blocks = data_list[start_idx:end_idx]
        
        # Create blocks for this sheet
        for local_idx, data in enumerate(sheet_blocks):
            # Calculate position in grid
            grid_col = local_idx % cols_per_row
            grid_row = local_idx // cols_per_row
            
            # Calculate actual Excel position
            start_col = 1 + (grid_col * block_width)
            start_row = 1 + (grid_row * block_height)
            
            # Create the block
            create_summary_block(ws, data, start_row, start_col)
        
        # Auto-adjust column widths for this sheet
        for col in range(1, cols_per_row * block_width + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 15
    
    # Save the file
    wb.save(filename)
    print(f"Excel file saved as: {filename}")
    print(f"Created {total_blocks} summary blocks across {num_sheets} sheet(s)")
    print(f"Blocks per sheet: {blocks_per_sheet} ({cols_per_row} columns)")

def create_excel_summary_single_sheet(df, filename="summary_layout_single_sheet.xlsx", cols_per_row=3):
    """
    Create Excel file with ALL summary blocks on a single sheet (for smaller datasets or if you prefer one sheet)
    """
    # Convert dataframe to block format
    data_list = process_dataframe_to_blocks(df)
    total_blocks = len(data_list)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "All_Summaries"
    
    # Configuration
    block_width = 4  # columns per block (3 data + 1 spacing)
    block_height = 7  # rows per block
    
    # Create all blocks on single sheet
    for idx, data in enumerate(data_list):
        # Calculate position in grid
        grid_col = idx % cols_per_row
        grid_row = idx // cols_per_row
        
        # Calculate actual Excel position
        start_col = 1 + (grid_col * block_width)
        start_row = 1 + (grid_row * block_height)
        
        # Create the block
        create_summary_block(ws, data, start_row, start_col)
    
    # Auto-adjust column widths
    for col in range(1, cols_per_row * block_width + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 15
    
    # Save the file
    wb.save(filename)
    print(f"Excel file saved as: {filename}")
    print(f"Created {total_blocks} summary blocks on single sheet")
    print(f"Sheet dimensions: {cols_per_row} columns x {(total_blocks + cols_per_row - 1) // cols_per_row} rows of blocks")


if __name__ == "__main__":
    import sys
    try:
        input_pickle_file = sys.argv[1]
        output_prefix = sys.argv[2]
    except IndexError:
        import rich
        rich.print("\n[bold red]I suspect you forgot to specify the file to read?[/bold red]\n **EXITING !!**[/bold red]\n"); sys.exit(1)
    import pandas as pd
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    df = pd.read_pickle(input_pickle_file)
    multi_sheet_fn = make_multi_sheet_output_fn(output_prefix) 
    single_sheet_fn = make_single_sheet_output_fn(output_prefix)
    create_excel_summary_from_df(df, multi_sheet_fn)
    create_excel_summary_single_sheet(df, single_sheet_fn)